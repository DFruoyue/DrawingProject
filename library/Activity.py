import random, os, sys, copy
from datetime import datetime
from openpyxl import Workbook, load_workbook
from .Member import *
from .GuaranteedPool import GuaranteedPool
from .GUIcomponents import output

class Activity(object):
    _enrolled_members:set[EnrolledMember] = set()
    _drawn_members:set[EnrolledMember] = set()

    def __init__(self, activityTime:str, N:int, applicantsFile:str, baseDir:str) -> None:
        self.__activityTime = activityTime
        self.__N = N
        self.__base_direction = baseDir
        self._enrolled_members.clear()
        self._drawn_members.clear()
        self._applicantsFile = self.__activityTime + '活动名单.xlsx'
        self.__load_enrolled_members(applicantsFile)


    def __load_enrolled_members(self, applicantsFile:str) -> None:
        if applicantsFile.endswith('.xlsx'):
            self.__load_enrolled_members_excel(applicantsFile)
        elif applicantsFile.endswith('.txt'):
            self.__load_enrolled_members_txt(applicantsFile)
        else:
            output(f"申请人文件错误,错误信息:[非法文件类型:{applicantsFile}]")
            sys.exit()


    def __load_enrolled_members_txt(self, applicants_file:str) -> None:
        try:
            self._enrolled_members.clear()
            with open(applicants_file, 'r', encoding='utf-8') as file:
                lines = file.read().splitlines()
                for line in lines:
                    m = line.split('\t')
                    if not m[0] == None:
                        applicant = EnrolledMember(name=str(m[1]), wechat_id=str(m[0]), group=str(m[2]), email_address=str(m[3]))
                        self._enrolled_members.add(applicant)
            return True
        except FileNotFoundError:
            output(f"申请人文件错误,错误信息:[未找到'{applicants_file}'文件]")
            return False
        except Exception as e:
            output(f"申请人文件错误,错误信息:[未预测到的错误:{e}]")
            return False


    def __load_enrolled_members_excel(self, applicants_file:str) -> None:
        try:
            self._enrolled_members.clear()
            wb = load_workbook(applicants_file)
            sheet = wb.active
            wechatId_idx = -1
            name_idx = -1
            group_idx = -1
            email_idx = -1
            for cell in sheet[1]:
                if '微信' in cell.value:
                    wechatId_idx = cell.column - 1
                elif '名字' in cell.value or '姓名' in cell.value:
                    name_idx = cell.column - 1
                elif '群' in cell.value or '群来源' in cell.value:
                    group_idx = cell.column - 1
                elif '邮箱' in cell.value:
                    email_idx = cell.column - 1
                else:
                    continue

            for row in sheet.iter_rows(min_row=2, values_only=True):
                applicant = EnrolledMember(name=str(row[name_idx]), wechat_id=str(row[wechatId_idx]), group_name=str(row[group_idx]), email_address=str(row[email_idx]))
                self._enrolled_members.add(applicant)

        except FileNotFoundError:
            output(f"申请人文件错误,错误信息:[未找到'{applicants_file}'文件]")
        except Exception as e:
            output(f"申请人文件错误,错误信息:[未预测到的错误:{e}]")


    def __saveData(self) ->str:#输出excel文件
        wb = Workbook()
        sheet = wb.active
        sheet['A1'] = self.__activityTime+'中签名单'
        sheet['B1'] = '会员群'
        sheet['C1'] = '邮箱'
        sheet['D1'] = '微信号'
        sheet['E1'] = '保底'
        for index, member in enumerate(self._drawn_members, start=2):
            sheet[f'A{index}'] = member.getName()
            sheet[f'B{index}'] = member.getGroupName()
            sheet[f'C{index}'] = member.getEmailAddress()
            sheet[f'D{index}'] = member.getWechatID()
            sheet[f'E{index}'] = '✓' if member.isGuaranteed() else ''
        #确保ActivityRecord文件夹存在
        wb_dir_path = os.path.normpath(os.path.join(self.__base_direction, "ActivityRecord"))
        if not os.path.exists(wb_dir_path):
            os.makedirs(wb_dir_path)
        wb_file_path = os.path.normpath(os.path.join(wb_dir_path, self._applicantsFile))
        wb.save(wb_file_path)
        return wb_file_path

    
    def draw(self, GuaranteeEnable:bool = True) -> str:
        if GuaranteeEnable:
            guaranteed_pool = GuaranteedPool(baseDir = self.__base_direction, activityTime = self.__activityTime)
            self._drawn_members = guaranteed_pool.getPreviligedMembers(self._enrolled_members, Nbound = self.__N)
            self._enrolled_members = self._enrolled_members - self._drawn_members
            self._drawn_members.update(random.sample(list(self._enrolled_members), self.__N - len(self._drawn_members)))
            self._enrolled_members = self._enrolled_members - self._drawn_members
            guaranteed_pool.update(drawnMembers = self._drawn_members, undrawnMembers = self._enrolled_members)
            guaranteed_pool.saveData()
        else:
            self._drawn_members = set(random.sample(list(self._enrolled_members), self.__N))
        return self.__saveData()