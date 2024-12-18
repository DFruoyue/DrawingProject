import smtplib, sys, json
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import load_workbook
from .Member import EnrolledMember
from .GUIcomponents import output

class Email(object):
    def __init__(self, emailConfigFile:str, paymentLink:str, paymentDeadline:str, activitySubject:str, contactName:str, contactWehcatID:str) -> None:
        self._payment_link = paymentLink
        self._payment_deadline = paymentDeadline
        self._activitySubject = activitySubject
        self.__contact_name = contactName
        self.__contact_wechatID = contactWehcatID
        if not self.__load_email_config(emailConfigFile):
            sys.exit()


    def __check_login(self)->bool:
        try:
            server = smtplib.SMTP(self.__smtp_server, self.__smtp_port)
            server.starttls()
            server.login(self.__sender_address, self.__sender_password)
            server.quit()
            return True
        except Exception as e:
            return False

    #加载邮箱配置
    def __load_email_config(self, emailConfigFile: str) -> bool:
        try:
            with open(emailConfigFile, 'r', encoding='utf-8') as file:
                config_data = json.load(file)
                self.__smtp_server = config_data.get('smtp_server', None)
                self.__smtp_port = config_data.get('smtp_port', None)
                self.__sender_address = config_data.get('sender_address', None)
                self.__sender_password = config_data.get('sender_password', None)
                if not self.__smtp_server:
                    output('SMTP服务器地址smtp_server未配置')
                    return False
                if not self.__smtp_port:
                    output('SMTP服务器端口smtp_port未配置')
                    return False
                if not self.__sender_address:
                    output('发件人地址sender_address未配置')
                    return False
                if not self.__sender_password:
                    output('发件人密码sender_password未配置')
                    return False
                if not self.__check_login():
                    output('发件人登录失败')
                    return False
                return True
        except FileNotFoundError:
            output(f"邮箱设置文件错误:[未找到'{emailConfigFile}'文件]")
            return False
        except json.JSONDecodeError:
            output(f"邮箱设置文件错误:['{emailConfigFile}'文件内容违反Json规则]")
            return False
        except Exception as e:
            output(f"邮箱设置文件错误:[未预测到的错误:{e}]")
            return False


    def __send_email(self, receiver:EnrolledMember) -> None:
        # 创建邮件内容
        message = MIMEMultipart()
        message['From'] = self.__sender_address
        message['To'] = receiver.getEmailAddress()
        message['Subject'] = self._activitySubject+"活动通知"

        # 邮件正文内容
        mail_content = f'''
            <!DOCTYPE html>
            <html lang="zh-cn">
            <head>
                <meta charset="UTF-8">
                <title>活动通知</title>
                <style>
                    .content {{
                        font-family: 'times new roman', times, serif;
                        font-size: 14pt;
                        color: #000000;
                        
                    }}
                    .highlight {{
                        color: #ff1f00;
                    }}
                    .indent {{
                        text-indent: 2em;
                    }}
                </style>
            </head>
            <body>
                <div class="content">亲爱的{receiver.getName()}同学：<br>你好！</div>
                <div class="content indent">
                    恭喜你抽中本次下厨房活动！
                </div>
                <div class="content indent">
                    请点开缴费问卷链接填写个人信息，然后扫码缴费并提交截图。
                </div>
                <div class="content indent">
                    提交后问卷会<strong>弹出活动群二维码</strong>，请扫码进群。注意：若<strong>{self._payment_deadline}</strong>前未提交问卷，视为<strong>&nbsp;<span class="highlight">放弃名额&nbsp;</span></strong>！
                </div>
                <div class="content indent">
                    如果未能及时扫码入群，请联系{self.__contact_name}同学（微信号：{self.__contact_wechatID}）
                </div>
                <div class="content">附缴费问卷链接：{self._payment_link}</div>
                <script src="https://cdn.jsdelivr.net/gh/highlightjs/cdn-release@11.5.1/build/highlight.min.js" type="text/javascript"></script>
                <script type="text/javascript">hljs.highlightAll();</script>
            </body>
            </html>
        '''
        message.attach(MIMEText(mail_content, 'html', 'utf-8'))

        try:
            # 连接到SMTP服务器
            session = smtplib.SMTP(self.__smtp_server, self.__smtp_port)
            session.starttls() # 启动TLS加密
            session.login(self.__sender_address, self.__sender_password) # 登录到SMTP服务器
            text = message.as_string()
            session.sendmail(self.__sender_address, receiver.getEmailAddress(), text)
            session.quit()

        except Exception as e:
            output(f"{receiver.__emailAddress}邮件发送失败:{e}")


    def send(self, participantsExcelFile:str) -> None:
        wb = load_workbook(participantsExcelFile)
        sheet = wb.active
        wechaID_idx = -1
        name_idx = -1
        group_idx = -1
        email_idx = -1

        for cell in sheet[1]:
            if '微信号' in cell.value:
                wechaID_idx = cell.column - 1
            elif '名单' in cell.value or '姓名' in cell.value:
                name_idx = cell.column - 1
            elif '群' in cell.value:
                group_idx = cell.column - 1
            elif '邮箱' in cell.value:
                email_idx = cell.column - 1
            else:
                continue

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                receiver = EnrolledMember(name=str(row[name_idx]), 
                                        wechat_id=str(row[wechaID_idx]), 
                                        group_name=str(row[group_idx]), 
                                        email_address=str(row[email_idx]))
                self.__send_email(receiver)
            except Exception as e:
                print(f"\t发送邮件给 {receiver.getName()}, {receiver.getEmailAddress()} | {row[email_idx]} 时出错: {e}")
        
        sheet.cell(row = sheet.max_row+1, column = 1, value = '活动群联系人:'+ self.__contact_name+ "(" + self.__contact_wechatID + ")")
        sheet.merge_cells(start_row = sheet.max_row, start_column = 1, end_row = sheet.max_row, end_column = 2)
        sheet.cell(row = sheet.max_row, column=3, value="已发送邮件")
        wb.save(participantsExcelFile)